# -*- coding: utf-8 -*-
"""Erzeugt Teilnehmer- und Lösungsarbeitsmappe aus der Satzspezifikation.

Die Lösungsdatei enthält echte Formeln in echten Zellen (§18.2), die
Teilnehmerdatei dieselbe Struktur mit leeren Zielbereichen (§6.1).
openpyxl schreibt Formeln in englischer Notation mit Komma als Trennzeichen;
Excel zeigt sie in deutscher Oberfläche mit den deutschen Namen an.
"""
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
F_TXT = Font(name=ARIAL, size=11)
F_BOLD = Font(name=ARIAL, size=11, bold=True)
_THIN = Side(style="thin", color="808080")
BORD = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
KOPFFARBE = PatternFill("solid", fgColor="D9D9D9")

# Zahlenformate nach §5.7
FMT = {
    "eur": '#,##0.00" €"',
    "prozent": '0.0%',
    "menge": '#,##0',
    "ganz": '0',
    "datum": 'DD.MM.YYYY',
    "text": '@',
    "dezimal": '#,##0.00',
}

DIAGRAMMTYP = {
    "balken": (BarChart, {"type": "bar"}),
    "saeule": (BarChart, {"type": "col"}),
    "linie": (LineChart, {}),
    "kreis": (PieChart, {}),
    "punkt": (ScatterChart, {}),
}


def _kopf_stylen(ws, zeile, anzahl, hoehe=40):
    for c in range(1, anzahl + 1):
        cell = ws.cell(row=zeile, column=c)
        cell.font = F_BOLD
        cell.border = BORD
        cell.fill = KOPFFARBE
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[zeile].height = hoehe


def _wert_schreiben(ws, zeile, spalte, wert, fmt=None, ausrichtung=None):
    cell = ws.cell(row=zeile, column=spalte, value=wert)
    cell.border = BORD
    cell.font = F_TXT
    if fmt:
        cell.number_format = FMT.get(fmt, fmt)
    if ausrichtung:
        cell.alignment = Alignment(horizontal=ausrichtung)
    return cell


def _kopfbreite_noetig(kopf, zeichenreserve=2.4):
    """Breite, bei der die Überschrift nie mitten im Wort umbricht (§18.3).

    Excel bricht in einer zu schmalen Zelle zeichenweise um — aus
    "Rabattsatz" wird "Rabattsat" und "z". Das verhindert nur eine Spalte,
    die mindestens so breit ist wie ihr längstes unteilbares Wort; zwischen
    Wörtern darf und soll dagegen umgebrochen werden.

    Gemessen wird in Excel-Zeichenbreiten, der Einheit von
    column_dimensions.width — nicht in Zentimetern.
    """
    woerter = str(kopf).replace("-", "- ").replace("/", "/ ").split()
    laengstes = max((len(w) for w in woerter), default=0)
    return laengstes + zeichenreserve


def _breite(kopf, vorgabe, standard=14):
    """Vorgegebene Breite, angehoben auf das Nötige für die Überschrift.

    So bleibt die Vorgabe aus der Satzspezifikation wirksam und der
    Zeichenumbruch ist trotzdem konstruktiv ausgeschlossen statt nur
    bemessen.
    """
    return max(vorgabe if vorgabe else standard, _kopfbreite_noetig(kopf))


def _datenblatt(wb, blatt):
    """Ein gefülltes Blatt: Stammdaten oder Zusatzblatt (§6.1)."""
    ws = wb.create_sheet(blatt["name"])
    ws["A1"] = blatt.get("titel", blatt["name"])
    ws["A1"].font = F_BOLD
    kopf = blatt["kopf"]
    kopfzeile = blatt.get("kopfzeile", 3)
    for i, t in enumerate(kopf, start=1):
        ws.cell(row=kopfzeile, column=i, value=t)
        vorgabe = blatt.get("breiten", [None] * len(kopf))[i - 1] \
            or max(12, min(30, len(str(t)) + 6))
        ws.column_dimensions[get_column_letter(i)].width = _breite(t, vorgabe)
    _kopf_stylen(ws, kopfzeile, len(kopf))
    formate = blatt.get("formate", [None] * len(kopf))
    for ri, row in enumerate(blatt["zeilen"]):
        z = kopfzeile + 1 + ri
        for ci, v in enumerate(row, start=1):
            fmt = formate[ci - 1]
            if fmt is None:
                if isinstance(v, dt.date):
                    fmt = "datum"
                elif isinstance(v, (int, float)):
                    fmt = "menge"
            aus = "left" if isinstance(v, str) else "right"
            _wert_schreiben(ws, z, ci, v, fmt, aus)
    return ws


def _auswertungsblatt(wb, blatt, loesung):
    """Das Blatt, in dem gerechnet wird.

    Jede Spalte ist entweder 'eingabe' (von den Teilnehmenden aus einer
    Anlage zu übernehmen) oder 'formel' (zu berechnen). In der
    Teilnehmerdatei bleiben beide leer, in der Lösungsdatei stehen Werte
    beziehungsweise Formeln.
    """
    ws = wb.create_sheet(blatt["name"])
    ws["A1"] = blatt.get("titel", blatt["name"])
    ws["A1"].font = F_BOLD
    spalten = blatt["spalten"]
    kopfzeile = blatt.get("kopfzeile", 3)
    erste = kopfzeile + 1
    anzahl = blatt["anzahl_zeilen"]

    for i, sp in enumerate(spalten, start=1):
        ws.cell(row=kopfzeile, column=i, value=sp["kopf"])
        ws.column_dimensions[get_column_letter(i)].width = \
            _breite(sp["kopf"], sp.get("breite"))
    _kopf_stylen(ws, kopfzeile, len(spalten))

    for ri in range(anzahl):
        z = erste + ri
        for ci, sp in enumerate(spalten, start=1):
            cell = ws.cell(row=z, column=ci)
            cell.border = BORD
            cell.font = F_TXT
            cell.alignment = Alignment(
                horizontal=sp.get("ausrichtung",
                                  "left" if sp.get("art") == "text" else "right"))
            if not loesung:
                continue
            if "formel" in sp:
                cell.value = sp["formel"].format(z=z, erste=erste,
                                                 letzte=erste + anzahl - 1)
            elif "werte" in sp:
                cell.value = sp["werte"][ri]
            if sp.get("format"):
                cell.number_format = FMT.get(sp["format"], sp["format"])

    for zusatz in blatt.get("einzelzellen", []):
        if zusatz.get("label"):
            lz = ws[zusatz["label_zelle"]]
            lz.value = zusatz["label"]
            lz.font = F_BOLD
        cell = ws[zusatz["zelle"]]
        cell.border = BORD
        cell.font = F_TXT
        cell.alignment = Alignment(horizontal="right")
        if loesung:
            cell.value = zusatz["formel"].format(erste=erste,
                                                 letzte=erste + anzahl - 1)
            if zusatz.get("format"):
                cell.number_format = FMT.get(zusatz["format"], zusatz["format"])

    dia = blatt.get("diagramm")
    if dia and loesung:
        klasse, kwargs = DIAGRAMMTYP[dia.get("typ", "saeule")]
        ch = klasse()
        for k, v in kwargs.items():
            setattr(ch, k, v)
        ch.title = dia["titel"]
        if hasattr(ch, "y_axis"):
            ch.y_axis.title = dia.get("wertachse", "")
            ch.x_axis.title = dia.get("rubrikachse", "")
        werte = Reference(ws, min_col=dia["wertespalte"], min_row=kopfzeile,
                          max_row=erste + anzahl - 1)
        rubriken = Reference(ws, min_col=dia["rubrikspalte"], min_row=erste,
                             max_row=erste + anzahl - 1)
        ch.add_data(werte, titles_from_data=True)
        ch.set_categories(rubriken)
        ch.height = dia.get("hoehe", 9)
        ch.width = dia.get("breite", 19)
        ws.add_chart(ch, dia.get("position", "A12"))
    return ws


def mappe_bauen(spec, loesung, pfad):
    """Baut eine Arbeitsmappe in der Blattreihenfolge der Spezifikation."""
    wb = Workbook()
    wb.remove(wb.active)
    for blatt in spec["blaetter"]:
        if blatt.get("art") == "auswertung":
            _auswertungsblatt(wb, blatt, loesung)
        else:
            _datenblatt(wb, blatt)
    aktiv = spec.get("aktives_blatt", spec["blaetter"][0]["name"])
    wb.active = wb.sheetnames.index(aktiv)
    wb.save(pfad)
    return pfad
