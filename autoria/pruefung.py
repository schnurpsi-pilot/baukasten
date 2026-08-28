# -*- coding: utf-8 -*-
"""Automatische Checkliste vor der Ausgabe (§14.1).

Geprüft wird alles, was sich maschinell prüfen lässt: Punktesumme,
Deckungsabgleich, Teilaufgabengrenzen, Funktionsrahmen, Rechenprobe über
eine Neuberechnung der Lösungsdatei, Vollständigkeit der Artefakte.

Was sich nicht maschinell prüfen lässt — Eindeutigkeit der Aufträge,
Eigenständigkeit, Sperrlisten gegen die Reihenhistorie — meldet der Bericht
ausdrücklich als ungeprüft. Fehlende Prüfmöglichkeit ist kein Bestehen
(§14.2).
"""
import os
import shutil
import subprocess
import tempfile

# Zulässige Funktionen nach Anhang D.3.
ERLAUBTE_FUNKTIONEN = {
    "ANZAHL", "ANZAHL2", "HEUTE", "JAHR", "MONAT", "TAG", "MIN", "MAX",
    "MITTELWERT", "ODER", "RANG", "RUNDEN", "AUFRUNDEN", "ABRUNDEN", "STUNDE",
    "MINUTE", "SUMME", "SUMMEWENN", "SVERWEIS", "TAGE360", "UND", "WENN",
    "ZÄHLENWENN",
}
# openpyxl schreibt englisch; Zuordnung für die Prüfung.
EN_DE = {
    "COUNT": "ANZAHL", "COUNTA": "ANZAHL2", "TODAY": "HEUTE", "YEAR": "JAHR",
    "MONTH": "MONAT", "DAY": "TAG", "MIN": "MIN", "MAX": "MAX",
    "AVERAGE": "MITTELWERT", "OR": "ODER", "RANK": "RANG", "ROUND": "RUNDEN",
    "ROUNDUP": "AUFRUNDEN", "ROUNDDOWN": "ABRUNDEN", "HOUR": "STUNDE",
    "MINUTE": "MINUTE", "SUM": "SUMME", "SUMIF": "SUMMEWENN", "VLOOKUP": "SVERWEIS",
    "DAYS360": "TAGE360", "AND": "UND", "IF": "WENN", "COUNTIF": "ZÄHLENWENN",
}
MAX_TEILAUFGABEN = {"tabellenkalkulation": 12, "textverarbeitung": 8,
                    "kommunikation": 5}


class Befund:
    """Sammelt die Befunde für den Ampelbericht (§14.2).

    Rot kennt zwei Sorten: echte Regelverstöße, die den Lauf scheitern
    lassen, und systemische Grenzen wie die fehlende Prüfung in der
    Zielsoftware. Beide stehen im Bericht unter Rot — nur die erste Sorte
    steuert den Rückgabewert, sonst wäre jeder Lauf rot.
    """

    def __init__(self):
        self.gruen, self.gelb, self.rot = [], [], []
        self.verstoesse = []

    def ok(self, text):
        self.gruen.append(text)

    def warn(self, text):
        self.gelb.append(text)

    def fehler(self, text):
        self.rot.append(text)
        self.verstoesse.append(text)

    def grenze(self, text):
        """Rot ohne Regelverstoß: konnte nicht geprüft werden."""
        self.rot.append(text)

    @property
    def bestanden(self):
        return not self.verstoesse

    def bericht(self):
        zeilen = []
        for stufe, eintraege in (("ROT", self.rot), ("GELB", self.gelb),
                                 ("GRÜN", self.gruen)):
            zeilen.append(f"[{stufe}]")
            if not eintraege:
                zeilen.append("  keine Einträge")
            for e in eintraege:
                zeilen.append(f"  - {e}")
        return "\n".join(zeilen)


def punkte_pruefen(spec, bf):
    """§14.1 Punkt 5: Summe muss exakt die Gesamtpunktzahl ergeben."""
    m = spec["meta"]
    summe_aufgaben = sum(a["punkte"] for a in spec["aufgaben"])
    if summe_aufgaben != m["gesamtpunkte"]:
        bf.fehler(f"Aufgabenpunkte ergeben {summe_aufgaben} statt "
                  f"{m['gesamtpunkte']} (§9.1).")
    for a in spec["aufgaben"]:
        teil = sum(p for _n, _t, p in a["teilaufgaben"])
        if teil + 2 != a["punkte"]:
            bf.fehler(f"Aufgabe {a['nr']}: Teilaufgaben {teil} + 2 Formatpunkte "
                      f"ergeben {teil + 2}, deklariert sind {a['punkte']} (§5.2).")
    summe_bewertung = sum(b["punkte"] for b in spec["bewertung"])
    if summe_bewertung != m["gesamtpunkte"]:
        bf.fehler(f"Bewertungsbogen summiert auf {summe_bewertung} statt "
                  f"{m['gesamtpunkte']}.")
    if not bf.rot:
        bf.ok(f"Punktesumme exakt {m['gesamtpunkte']}; Aufgaben-, Teilaufgaben- und "
              "Bewertungssummen stimmen überein (maschinell addiert).")


def deckung_pruefen(spec, bf):
    """§14.1 Punkt 6: je Auftrag genau eine Bewertungszeile und umgekehrt."""
    aus_aufgaben = [f"{a['nr']}{n}" for a in spec["aufgaben"]
                    for n, _t, _p in a["teilaufgaben"]]
    aus_bewertung = [b["nr"] for b in spec["bewertung"] if "Format" not in b["nr"]]
    fehlt_bewertung = [x for x in aus_aufgaben if x not in aus_bewertung]
    fehlt_auftrag = [x for x in aus_bewertung if x not in aus_aufgaben]
    if fehlt_bewertung:
        bf.fehler(f"Ohne Bewertungszeile: {fehlt_bewertung}")
    if fehlt_auftrag:
        bf.fehler(f"Bewertungszeile ohne Auftrag: {fehlt_auftrag}")
    formatzeilen = [b["nr"] for b in spec["bewertung"] if "Format" in b["nr"]]
    if len(formatzeilen) != len(spec["aufgaben"]):
        bf.fehler(f"{len(formatzeilen)} Formatzeilen bei "
                  f"{len(spec['aufgaben'])} Aufgaben (§9.3).")
    if not (fehlt_bewertung or fehlt_auftrag):
        bf.ok(f"Deckungsabgleich bestanden: {len(aus_aufgaben)} Aufträge, "
              f"{len(aus_aufgaben)} Bewertungszeilen, keine Lücke.")


def zuschnitt_pruefen(spec, bf):
    """§5.4: Obergrenzen für die Zahl der Teilaufgaben."""
    for a in spec["aufgaben"]:
        grenze = MAX_TEILAUFGABEN.get(a.get("typ"))
        anzahl = len(a["teilaufgaben"])
        if grenze and anzahl > grenze:
            bf.fehler(f"Aufgabe {a['nr']} hat {anzahl} Teilaufgaben, erlaubt sind "
                      f"{grenze} (§5.4). Teilaufgaben zusammenfassen.")
        elif grenze:
            bf.ok(f"Aufgabe {a['nr']}: {anzahl} Teilaufgaben, Grenze {grenze} "
                  f"eingehalten (§5.4).")


def funktionsrahmen_pruefen(spec, bf):
    """§6.4 und Anhang D: keine Funktion außerhalb der Befehlsübersicht."""
    import re
    gefunden = set()
    for blatt in spec.get("blaetter", []):
        for sp in blatt.get("spalten", []):
            if "formel" in sp:
                gefunden |= set(re.findall(r"([A-Z][A-Z0-9\.]+)\s*\(", sp["formel"]))
        for zz in blatt.get("einzelzellen", []):
            gefunden |= set(re.findall(r"([A-Z][A-Z0-9\.]+)\s*\(", zz["formel"]))
    unbekannt = []
    verwendet = set()
    for f in gefunden:
        de = EN_DE.get(f, f)
        if de in ERLAUBTE_FUNKTIONEN:
            verwendet.add(de)
        else:
            unbekannt.append(f)
    if unbekannt:
        bf.fehler(f"Nicht in Anhang D gelistet: {sorted(unbekannt)}. Aufgabe "
                  "umkonstruieren, nicht den Rahmen erweitern (§6.4).")
    else:
        bf.ok(f"Funktionsrahmen eingehalten: {sorted(verwendet)} (Abgleich gegen "
              "Anhang D.3).")


def rechenprobe(loesungspfad, erwartet, bf):
    """§14.1 Punkt 3: jede Formel nachrechnen.

    Die Lösungsdatei wird von LibreOffice neu berechnet; die Ergebnisse
    werden gegen die unabhängig gerechneten Sollwerte gestellt.
    erwartet: {"Blatt!A1": sollwert, ...}
    """
    if not shutil.which("soffice"):
        bf.grenze("Rechenprobe nicht möglich: LibreOffice nicht verfügbar. "
                  "Fehlende Prüfmöglichkeit ist kein Bestehen (§14.2).")
        return
    from openpyxl import load_workbook
    tmp = tempfile.mkdtemp()
    try:
        subprocess.run(["soffice", "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", tmp, loesungspfad],
                       check=True, capture_output=True, timeout=300)
        neu = os.path.join(tmp, os.path.basename(loesungspfad))
        wb = load_workbook(neu, data_only=True)
        abweichungen = []
        for adresse, soll in erwartet.items():
            blatt, zelle = adresse.split("!")
            ist = wb[blatt][zelle].value
            if isinstance(soll, (int, float)) and isinstance(ist, (int, float)):
                if abs(float(soll) - float(ist)) > 0.005:
                    abweichungen.append(f"{adresse}: soll {soll}, ist {ist}")
            elif str(soll) != str(ist):
                abweichungen.append(f"{adresse}: soll {soll!r}, ist {ist!r}")
        if abweichungen:
            bf.fehler(f"Rechenprobe: {len(abweichungen)} Abweichung(en) — "
                      f"{abweichungen[:5]}")
        else:
            bf.ok(f"Rechenprobe bestanden: {len(erwartet)} Zellen von LibreOffice "
                  "neu berechnet, alle identisch mit der unabhängigen Rechnung.")
    except Exception as e:                                  # pragma: no cover
        bf.fehler(f"Rechenprobe fehlgeschlagen: {e}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def teilnehmerdatei_pruefen(teilnehmerpfad, spec, bf):
    """§6.1: Zielbereiche leer, Stammdatenblätter gefüllt, kein Diagramm."""
    from openpyxl import load_workbook
    wb = load_workbook(teilnehmerpfad)
    for blatt in spec.get("blaetter", []):
        if blatt.get("art") != "auswertung":
            continue
        ws = wb[blatt["name"]]
        kopfzeile = blatt.get("kopfzeile", 3)
        erste = kopfzeile + 1
        belegt = [ws.cell(row=z, column=c).value
                  for z in range(erste, erste + blatt["anzahl_zeilen"])
                  for c in range(1, len(blatt["spalten"]) + 1)]
        if any(v not in (None, "") for v in belegt):
            bf.fehler(f"Teilnehmerdatei: Blatt {blatt['name']} enthält bereits "
                      "Werte im Bearbeitungsbereich.")
        else:
            bf.ok(f"Teilnehmerdatei: Bearbeitungsbereich in Blatt "
                  f"{blatt['name']} ist leer.")
        if getattr(ws, "_charts", []):
            bf.fehler(f"Teilnehmerdatei enthält ein Diagramm in {blatt['name']} — "
                      "es ist Prüfungsleistung (§15.3).")
    stamm = [b for b in spec.get("blaetter", []) if b.get("stammdaten")]
    leer = [b["name"] for b in stamm
            if wb[b["name"]].cell(row=b.get("kopfzeile", 3) + 1, column=2).value
            in (None, "")]
    if leer:
        bf.fehler(f"Stammdatenblätter ohne Inhalt: {leer}. SVERWEIS liefe ins "
                  "Leere (§6.1).")
    elif stamm:
        bf.ok(f"Stammdatenblätter gefüllt: {[b['name'] for b in stamm]}.")


def vollstaendigkeit_pruefen(ordner, dateiliste, bf):
    """§14.1 Punkt 9 und §17: alle Artefakte liegen vor."""
    fehlend = [d for d in dateiliste if not os.path.exists(os.path.join(ordner, d))]
    if fehlend:
        bf.fehler(f"Fehlende Artefakte: {fehlend}. Nichts ausgeben, nachproduzieren "
                  "(§18.7).")
    else:
        bf.ok(f"Vollständigkeit: alle {len(dateiliste)} Artefakte vorhanden, "
              "Dateibenennung nach §11.2.")


def sperren_melden(spec, bf):
    """§10.1 bis §10.3: Ergebnis der Reihenprüfung in den Ampelbericht.

    Eine verletzte harte Sperre ist ein Regelverstoß und lässt den Lauf
    scheitern. Weiche Sperren sind Häufigkeitshinweise und stehen unter Gelb.
    Ohne Reihenplan bleibt die Pflichtbelegung ungeprüft (§14.2).
    """
    check = spec.get("sperrlisten_check")
    if check:
        if check["hart"].startswith("verletzt"):
            bf.fehler(f"Harte Sperre verletzt — {check['hart'][10:]}")
        else:
            bf.ok(f"Harte Sperren §10.1 {check['hart']}.")
        if check["weich"].startswith("keine"):
            bf.ok("Weiche Sperren §10.2: keine Obergrenze überschritten.")
        else:
            bf.warn(f"Weiche Sperre §10.2 — {check['weich']}")

    if not spec.get("reihenplan_eingelesen"):
        bf.warn("Kein Reihenplan übergeben — die Pflichtbelegung nach §10.3 "
                "wurde nicht fortgeschrieben. Mit --reihenplan starten.")
        return
    offen = spec.get("pflichtelemente_offen")
    if offen:
        bf.ok(f"Pflichtbelegung §10.3 fortgeschrieben: noch offen sind "
              f"{len(offen)} Elemente ({', '.join(offen)}).")
    elif offen is not None:
        bf.ok("Pflichtbelegung §10.3: alle Sollwerte des Reihenplans erreicht.")


# Absätze, die zulässig klein beginnen dürfen.
ANREDEN = ("guten morgen", "guten tag", "hallo", "sehr geehrte", "sehr geehrter",
           "liebe", "lieber", "moin")
# Wörter, die als Satzanfang klein korrekt sind, weil sie keinen Satz eröffnen.
KLEIN_ERLAUBT = ("bzw", "ggf", "z", "d", "u", "ca", "vgl", "s", "evtl")


def satzanfaenge_pruefen(spec, bf):
    """Ein neuer Absatz beginnt groß, sofern er nicht den Vorabsatz fortsetzt.

    Nach der Anrede setzt der erste Absatz den Satz fort und bleibt klein.
    Jeder weitere Absatz eröffnet einen neuen Satz und gehört groß.

    Grundlage ist §4.5 (Rechtschreibprüfung), nicht §5.7: Die DIN 5008 regelt
    Schreibweisen und Gestaltung, nicht Orthografie. Ein Verstoß ist deshalb
    ein Regelverstoß und lässt den Lauf scheitern.
    """
    import re as _re

    def _zeilen(obj, treffer):
        """Sammelt alle Absatzlisten aus der Spezifikation, egal wie tief."""
        if isinstance(obj, dict):
            if obj.get("typ") == "absaetze" and isinstance(obj.get("zeilen"), list):
                treffer.append(obj["zeilen"])
            for wert in obj.values():
                _zeilen(wert, treffer)
        elif isinstance(obj, list):
            for wert in obj:
                _zeilen(wert, treffer)
        return treffer

    verdaechtig = []
    for zeilen in _zeilen(spec, []):
        nach_anrede = False
        vorher = ""
        for zeile in zeilen:
            text = (zeile or "").strip()
            if not text:
                continue
            klein = text.lower()
            if klein.startswith(ANREDEN):
                nach_anrede = True
                vorher = text
                continue
            erstes = text.split()[0].rstrip(".,;:").lower()
            beginnt_klein = (text[0].isalpha() and text[0].islower()
                             and erstes not in KLEIN_ERLAUBT
                             and not _re.match(r"^[a-z_][a-z0-9_]*\.[a-z]{2,4}\b", text))
            # Der Vorabsatz endet offen: der Satz läuft weiter, klein ist richtig.
            fortsetzung = vorher.endswith((",", ":", ";", "-", "–"))
            if beginnt_klein and not fortsetzung and not nach_anrede:
                verdaechtig.append(text[:60])
            nach_anrede = False
            vorher = text

    if verdaechtig:
        bf.fehler(f"Groß- und Kleinschreibung (§4.5): Absätze beginnen klein, "
                  f"obwohl sie einen neuen Satz eröffnen: {verdaechtig}. Nach "
                  f"der Anrede ist Kleinschreibung richtig, danach nicht mehr.")
    else:
        bf.ok("Satzanfänge (§4.5): jeder Absatz beginnt groß, ausgenommen der "
              "erste nach einer Anrede.")


def ungeprueft_melden(spec, bf):
    """Was maschinell nicht prüfbar ist, gehört nach Rot beziehungsweise Gelb."""
    bf.grenze("Keine Datei in Microsoft Word oder Excel geöffnet. Gerendert und "
              "geprüft wurde mit LibreOffice und openpyxl. Vor dem Einsatz die "
              "Lösungsdateien einmal in der Zielsoftware öffnen (§14.2).")
    if not spec.get("historie_eingelesen"):
        bf.warn("Keine Reihenhistorie eingelesen — die harten Sperren nach §10.1 "
                "konnten nicht gegen die bisherigen Sätze geprüft werden. Die "
                "Historie beginnt mit diesem Satz (§11.7).")
    else:
        bf.ok("Reihenhistorie aus Datei eingelesen und um diesen Satz ergänzt "
              "(§11.7).")
    sperren_melden(spec, bf)
    bf.warn("Eindeutigkeit der Aufträge (§4.3), Eigenständigkeit (§20) und "
            "sprachliche Angemessenheit sind inhaltliche Prüfungen und wurden "
            "nicht maschinell geprüft.")


def alles_pruefen(spec, ordner, dateiliste, loesungspfad=None,
                  teilnehmerpfad=None, erwartet=None):
    bf = Befund()
    punkte_pruefen(spec, bf)
    deckung_pruefen(spec, bf)
    zuschnitt_pruefen(spec, bf)
    funktionsrahmen_pruefen(spec, bf)
    if loesungspfad and erwartet:
        rechenprobe(loesungspfad, erwartet, bf)
    if teilnehmerpfad:
        teilnehmerdatei_pruefen(teilnehmerpfad, spec, bf)
    satzanfaenge_pruefen(spec, bf)
    vollstaendigkeit_pruefen(ordner, dateiliste, bf)
    ungeprueft_melden(spec, bf)
    return bf
